# make executable with pyinstaller -F 'file.py'

import os
import queue
import timeit
import shutil
from sys import exit as systemexit

import openpyxl as xl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, cols_from_range
from openpyxl.utils import get_column_letter
from openpyxl.comments.comment_sheet import Properties

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mb
from tkinter import filedialog
from PIL import Image, ImageTk

from copy import copy


def main():
    root = tk.Tk()
    root.title("AODA EXCEL FILE CONVERSION")
    beginApp(root)
    root.mainloop()


def appSize(root, h, w):
    appHeight = h
    appWidth = w
    # screenHeight = root.winfo_screenheight()
    screenWidth = root.winfo_screenwidth()
    xCoord = int((screenWidth / 2 - appWidth / 2))
    # yCoord = int((screenHeight/2 - appHeight/2))

    root.geometry("{}x{}+{}+{}".format(appWidth, appHeight, xCoord, 1))


def beginApp(root):
    imgInstructions = ''
    appHeight = 160
    appWidth = 1320
    appSize(root, appHeight, appWidth)

    # FRAME1 ELEMENTS
    frame1 = tk.LabelFrame(root, text="MAKING YOUR EXCEL FILE AODA COMPLIANT", width=1300, height=150, bd=5)

    img = ImageTk.PhotoImage(Image.open("mohLogoM.png").resize((171, 74), Image.ANTIALIAS))

    side = tk.Frame(frame1)
    imgLogo = tk.Label(side, image=img)
    imgLogo.img = img
    b1 = tk.Button(side, text="ABOUT", command=lambda: PopUp(root))

    selectFile = tk.Label(frame1, text="SELECT FILE:", font=(None, 12), width=12)
    selectedFileLabel = tk.Label(frame1, text="SELECTED FILE: ", font=(None, 12))
    selectedFileName = tk.Label(frame1, text="/path/filename ", font=(None, 12), fg='blue')
    messageLabel = tk.Label(frame1, text="MESSAGE: ", font=(None, 12))
    messageLabelText = tk.Label(frame1, text="HIT BROWSE TO PROCESS FILE.", font=(None, 12, 'bold'), fg='red')
    browseButton = tk.Button(frame1, text='BROWSE',
                             command=lambda: startProcess(root, browseButton, selectedFileName, messageLabelText),
                             width=20, font=(None, 12))

    # REMOVE START BUTTON AS PER USER FEEDBACK
    # startButton = tk.Button(frame1, text='START', state='disabled',
    #                         command=lambda: startAODA(root, selectedFileName, browseButton, startButton,
    #                                                   messageLabelText), width=20, font=(None, 12))
    # startButton.grid(row=1, column=1, padx=(2, 15), pady=(3, 3), sticky='ns')

    # FRAME1 GRID LAYOUT
    frame1.grid(row=0, column=0, columnspan=4, padx=10, pady=10)
    frame1.grid_propagate(False)

    side.grid(row=0, rowspan=2, column=0, padx=(2, 10), sticky='ns')
    imgLogo.grid(row=0, column=0, sticky='nw')
    b1.grid(row=1, padx=(5, 5), pady=(10, 0), sticky='ew')

    for x in range(2):
        frame1.grid_rowconfigure(x, weight=1)

    browseButton.grid(row=0, column=1, padx=(2, 15), pady=(3, 3), sticky='ns')
    selectedFileLabel.grid(row=0, column=2, padx=(0, 10), sticky='ew')
    selectedFileName.grid(row=0, column=3, sticky='w')

    messageLabel.grid(row=1, column=2, padx=(0, 10), sticky='ew')
    messageLabelText.grid(row=1, column=3, sticky='w')


def quitApp(wb, aodaFile):
    wb.save(aodaFile)
    wb.close()
    os.chmod(aodaFile, 0o755)
    systemexit()


def finishProcess(wb, aodaFile, browseButton):
    wb.save(aodaFile)
    wb.close()
    os.chmod(aodaFile, 0o755)
    print("PROCESS COMPLETE ... HIT BROWSE TO PROCESS ANOTHER FILE")
    browseButton["state"] = "normal"
    # startButton["state"] = "normal"


def isValid(cellAddress):
    try:
        return coordinate_from_string(cellAddress)
    except:
        # mb.showinfo("Alert", "Please enter valid cell address format in highlighted cells")
        pass


# TO COMBINE getFile and startAODA in one button as per USER FEEDBACK
def startProcess(root, browseButton, selectedFile, mlt):
    getFile(browseButton, selectedFile, mlt)
    startAODA(root, selectedFile, browseButton, mlt)


def getFile(browseButton, selectedFile, mlt):
    fileName = filedialog.askopenfilename()
    if fileName:
        selectedFile['text'] = fileName
        browseButton["state"] = "disabled"
        # startButton["state"] = "normal"
        mlt.configure(text='FILE OPENED.', font=(None, 12, 'bold'), fg='red')
    else:
        # exits if no file is chosen / cancels
        systemexit()


def startAODA(root, selectedFileName, browseButton, messageLabelText):
    # startButton["state"] = "disable"
    AODA(root, selectedFileName['text'], browseButton, messageLabelText)


class PopUp(tk.Toplevel):
    def __init__(self, root):
        tk.Toplevel.__init__(self, root)

        self.h = 40
        self.w = 95
        self.transient(root)
        self.grab_set()
        self.focus()
        self.UI()

    def UI(self):
        text = 'INFORMATION: ' \
               '\n' \
               '\n Version:\t\t1.0' \
               '\n Published:\tFebruary 2020' \
               '\n' \
               '\n Author:\t\tFalcon, JJ' \
               '\n Author:\t\tZhang, Alex' \
               '\n' \
               '\n INPUT:\tEXCEL FILE WITH TABLES' \
               '\n OUTPUT:\tEXCEL FILE FORMATTED TO BE AODA COMPLIANT' \
               '' \
               '\n \n FUNCTIONALITY:' \
               '\n \t - Translates all formulas to its value' \
               '\n \t - Un-merges merged cells' \
               '\n \t - Deletes any empty rows' \
               '\n \t - Updates header locations affected by deleted rows' \
               '\n \t - Updates un-merged cell locations affected by deleted rows' \
               '\n \t - Inserts "End of Column" appropriately' \
               '\n \t - Copies merged content across previous merged cell range' \
               '\n \t - Moves all appropriate column headers to the same row as title header' \
               '\n \t - Leaves blank cells inside Table as empty cells' \
               '\n \t - Converts Table into an "Excel Table Format"' \
               '\n \t - Formats Table with borders and alignment' \
               '\n \t - Updates "Screen Reader" text' \
               '\n \t - Writes "No Data" text on empty cells outside table boundaries' \
               '\n \t - TO DO: Define the table Header Names' \
               '\n' \
               '\n \n LIMITATIONS:' \
               '\n \t - Restricts application to sheets with <= 5 tables' \
               '\n \t - Cells inside the header segment that are blank will be left as-is' \
               '\n \t - Single blank cells above data segment will be used as headers' \
               '' \
               '\n \n USER INTERFACE DESCRIPTION' \
               '\n \t Navigation:' \
               '\n \t \t Convert\t\t : Triggers the conversion process' \
               '\n \t \t Prev\t\t : Allows user to go back to previous sheets for processing' \
               '\n \t \t Next\t\t : Allows user to skip ahead / continue with next sheet' \
               '\n \t \t Done\t\t : Allows user to process another file or quit the app' \
               '' \
               '\n \n UPDATES FROM VERSION 0.0' \
               '\n \t To process multiple tables in 1 sheet, that are not side by side' \
               '\n \t Names defined for each table' \
               '\n \n \n'

        self.title("About V1.0")

        self.msg = tk.Label(self, width=self.w, height=self.h, text=text, justify="left")
        self.close = tk.Button(self, text='OK', width=5, bg='#87CEFA', command=self.destroy)

        self.msg.grid(row=0, column=0, padx=(5, 10), pady=15, sticky='nw')
        self.close.grid(row=1, column=1, padx=20, pady=20, sticky='ew')


class AODA:
    def __init__(self, root, file, bb, mlt):
        self.root = root
        self.file = file
        self.aodaFile = self.file[0:self.file.rfind('.')] + '-AODA' + self.file[self.file.rfind('.'):len(self.file)]
        self.wb = ''
        self.sheetTotalCount = 0
        self.sheetCounter = 0
        self.sheetDicList = []
        self.tableList = []
        self.bb = bb
        # self.sb = sb
        self.messageLabelText = mlt
        self.imgInstructions = ''
        self.q = queue.Queue()
        self.copyFile()
        self.convertGUI()

    def copyFile(self):
        # create AODA copy of file
        shutil.copy(self.file, self.aodaFile)

        # open workbook
        self.wb = xl.load_workbook(self.aodaFile, data_only=True)
        self.wb.save(self.aodaFile)

    def getSheets(self):
        # counts number of sheets
        self.sheetTotalCount = len(self.wb.sheetnames)

        self.sheetDicList = [{'titleCell': '', 'sheetName': name, 'converted': False} for name in self.wb.sheetnames]

        self.sheetStatus = tk.StringVar()
        self.sheetStatus.set("SHEET %s OF %s" % (self.sheetCounter + 1, self.sheetTotalCount))

    def convertGUI(self):
        appHeight = 735
        appWidth = 1320
        appSize(self.root, appHeight, appWidth)

        self.getSheets()

        # FRAME2 ELEMENTS ###########################################################################################
        self.frame2 = tk.LabelFrame(self.root, text='SHEET INFO', width=750, height=470, bd=5)
        self.frame2.configure(text="SHEET DETAILS")

        titlebar2 = tk.Frame(self.frame2)

        self.sheetStatusButton = tk.Label(titlebar2, textvariable=self.sheetStatus, borderwidth=2, relief='raised',
                                          font=(None, 12), width=12, height=2)
        self.sheetName = tk.Label(titlebar2, text='', font=(None, 12, 'bold'), fg='blue', height=2)
        self.sheetName.configure(text=self.sheetDicList[self.sheetCounter]['sheetName'])

        sheetInstructionLabel = tk.Label(self.frame2, text="Enter Cell Address for the following:", font=(None, 12))
        reportTitleLabel = tk.Label(self.frame2, text="REPORT TITLE:", font=(None, 12))

        step1Label = tk.Label(self.frame2, text="STEP 1: ", font=(None, 12))
        step2Label = tk.Label(self.frame2, text="STEP 2: ", font=(None, 12))
        step3Label = tk.Label(self.frame2, text="STEP 3: ", font=(None, 12))

        startHeaderLabel = tk.Label(self.frame2, text="BEGINNING HEADER", font=(None, 12))
        startDataLabel = tk.Label(self.frame2, text="STARTING DATA SEGMENT", font=(None, 12))
        endDataLabel = tk.Label(self.frame2, text="ENDING DATA SEGMENT", font=(None, 12))

        self.submitButton = tk.Button(self.frame2, text="SUBMIT", command=self.generateTableList, width=15, height=2,
                                      font=(None, 12), bg='#87CEFA')

        # FRAME2 GRID
        self.frame2.grid(row=1, column=0, sticky='n', padx=(10, 0))
        self.frame2.grid_propagate(False)
        for x in range(1, 4):
            self.frame2.grid_columnconfigure(x, weight=1)

        titlebar2.grid(row=0, column=1, columnspan=3, padx=20, sticky='w')
        self.sheetStatusButton.grid(row=0, column=0, padx=5, pady=(15, 5))
        self.sheetName.grid(row=0, column=1, padx=5, pady=(15, 5), sticky='nw')

        sheetInstructionLabel.grid(row=1, column=1, padx=(22, 5), pady=(5, 10), sticky='nw')
        reportTitleLabel.grid(row=1, column=2, padx=(22, 5), pady=(5, 10))

        step1Label.grid(row=4, column=1, padx=5, pady=5)
        step2Label.grid(row=4, column=2, padx=5, pady=5)
        step3Label.grid(row=4, column=3, padx=5, pady=5)
        startHeaderLabel.grid(row=5, column=1, padx=5, pady=5)
        startDataLabel.grid(row=5, column=2, padx=5, pady=5)
        endDataLabel.grid(row=5, column=3, padx=5, pady=5)

        self.processEntryWidgets('SET')

        self.submitButton.grid(row=17, column=3, padx=5, pady=(15, 5))

        # FRAME3 ELEMENTS ###########################################################################################
        frame3 = tk.LabelFrame(self.root, text="INSTRUCTIONS", width=550, height=560, bd=5)
        frame3.grid_propagate(False)
        frame3.grid(row=1, column=1, rowspan=2, padx=(0, 10), sticky='nw')

        img = ImageTk.PhotoImage(Image.open("AODA.png").resize((525, 515), Image.ANTIALIAS))
        self.imgInstructions = img

        imgLabel = tk.Label(frame3, image=self.imgInstructions)
        imgLabel.grid(row=1, column=1, columnspan=5, padx=2, pady=2)

        # FRAME4 ELEMENTS ###########################################################################################
        frame4 = tk.LabelFrame(self.root, text="CONVERSION", width=750, height=90, bd=5)
        self.convertButton = tk.Button(frame4, text="CONVERT", state='disabled', command=self.convertSheet,
                                       height=3, font=(None, 12))
        self.prevButton = tk.Button(frame4, text=" PREV ", state='disabled', command=lambda: self.getNext(False),
                                    height=3, font=(None, 12))
        self.nextButton = tk.Button(frame4, text=" NEXT ", command=lambda: self.getNext(True), height=3,
                                    font=(None, 12))
        self.doneButton = tk.Button(frame4, text=" DONE ", command=self.closingProcess, height=3, font=(None, 12))

        frame4.grid_propagate(False)
        frame4.grid(row=2, column=0, sticky='w', padx=(10, 0))
        # frame4.grid_columnconfigure(0, minsize=100)
        for x in range(4):
            frame4.grid_columnconfigure(x, weight=1)

        self.convertButton.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.prevButton.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        self.nextButton.grid(row=1, column=2, padx=10, pady=5, sticky="ew")
        self.doneButton.grid(row=1, column=3, padx=10, pady=5, sticky="ew")

    def convertSheet(self):
        # Condition not required, submit/convert button should be disabled
        # but placed for safety net.  Checks in def getNext
        if not self.sheetDicList[self.sheetCounter]['converted']:
            self.buttonState('d', True, True, False, False)

            self.tableList = sorted(self.tableList, key=lambda k: (k['beginHeader'][0], k['beginHeader'][1]))

            print("startConversion: ", self.wb[self.sheetName['text']], self.tableList)

            print("TITLE ADDY: ", self.sheetDicList[self.sheetCounter]['titleHeader'])
            formatReportTitle(self.wb[self.sheetName['text']], self.sheetDicList[self.sheetCounter]['titleHeader'])

            startConversionProcess(self.wb[self.sheetName['text']], self.tableList)

            self.sheetDicList[self.sheetCounter]['converted'] = True
            self.messageLabelText.configure(text='SHEET CONVERTED.', font=(None, 12), fg='black')

            print("SAVING FILE ...")
            self.wb.save(self.aodaFile)
        else:
            self.messageLabelText.configure(text='sheet already CONVERTED', font=(None, 12, 'bold'), fg='red')

    def generateTableList(self):
        temp = []
        flag = flag1 = flag2 = flag3 = flag4 = flag5 = True

        if self.titleHeader.get():
            the = isValid(self.titleHeader.get())
            if the:
                self.titleHeader.configure(bg=self.origBG, fg=self.origFG)
                self.sheetDicList[self.sheetCounter]['titleHeader'] = the
            else:
                self.titleHeader.focus_set()
                self.titleHeader.configure(bg='#ffffe0', fg='red')
                flag = False

        if self.startHeaderEntry1.get() or self.startDataEntry1.get() or self.endDataEntry1.get():
            if self.getValidEntry('E1'):
                temp.append(self.getValidEntry('E1'))
            else:
                flag1 = False

        if self.startHeaderEntry2.get() or self.startDataEntry2.get() or self.endDataEntry2.get():
            if self.getValidEntry('E2'):
                temp.append(self.getValidEntry('E2'))
            else:
                flag2 = False

        if self.startHeaderEntry3.get() or self.startDataEntry3.get() or self.endDataEntry3.get():
            if self.getValidEntry('E3'):
                temp.append(self.getValidEntry('E3'))
            else:
                flag3 = False

        if self.startHeaderEntry4.get() or self.startDataEntry4.get() or self.endDataEntry4.get():
            if self.getValidEntry('E4'):
                temp.append(self.getValidEntry('E4'))
            else:
                flag4 = False

        if self.startHeaderEntry5.get() or self.startDataEntry5.get() or self.endDataEntry5.get():
            if self.getValidEntry('E5'):
                temp.append(self.getValidEntry('E5'))
            else:
                flag5 = False

        if flag and flag1 and flag2 and flag3 and flag4 and flag5:
            self.tableList = temp
            self.buttonState('n', False, True, False, False)
        else:
            mb.showinfo("Alert", "Please enter VALID CELL ADDRESS FORMAT")

    def getNext(self, forward):
        self.processEntryWidgets('CLEAR')
        self.messageLabelText.configure(text='ENTER SHEET DETAILS.  Click SUBMIT/CONVERT to start conversion.',
                                        font=(None, 12), fg='black')
        self.buttonState('d', False, True, False, False)
        self.buttonState('n', True, False, True, True)

        DicListSize = len(self.sheetDicList)

        if forward:
            self.sheetCounter += 1
        else:
            self.sheetCounter -= 1

        if self.sheetCounter == DicListSize:
            self.messageLabelText.configure(text='ONLY 1 SHEET TO PROCESS, click PREV to process it',
                                            font=(None, 12, 'bold'), fg='red')
            self.buttonState('d', True, True, False, True)
        else:
            if 0 <= self.sheetCounter < DicListSize:
                self.sheetStatus.set("SHEET %s OF %s" % (self.sheetCounter + 1, self.sheetTotalCount))
                self.sheetName.configure(text=self.sheetDicList[self.sheetCounter]['sheetName'])

                if self.sheetCounter == 0:
                    self.buttonState('d', False, True, True, False)

                elif self.sheetCounter == DicListSize - 1:
                    self.messageLabelText.configure(text='REACHED LAST ITEM', font=(None, 12, 'bold'), fg='red')
                    self.buttonState('d', False, True, False, True)

                if self.sheetDicList[self.sheetCounter]['converted']:
                    self.buttonState('d', True, False, False, False)
                    self.messageLabelText.configure(text='sheet already CONVERTED', font=(None, 12, 'bold'), fg='red')

    def closingProcess(self):
        finishProcess(self.wb, self.aodaFile, self.bb)
        self.buttonState('d', True, True, True, True)

        self.sheetStatus.set('GET SHEET')
        self.sheetName.configure(text='HIT BROWSE/START BUTTON')

        self.sheetTotalCount = 0
        self.sheetCounter = 0
        self.sheetDicList = []

        if mb.askyesno('PROCESS COMPLETE', 'Process Another File?'):
            finishProcess(self.wb, self.aodaFile, self.bb)
        else:
            quitApp(self.wb, self.aodaFile)

        self.messageLabelText.configure(text='PROCESS COMPLETE ... HIT BROWSE TO PROCESS ANOTHER FILE', font=(None, 12),
                                        fg='red')

    def clearEntry(self, she, sde, ede):
        she.delete(0, 'end')
        she.configure(bg=self.origBG, fg=self.origFG)
        sde.delete(0, 'end')
        sde.configure(bg=self.origBG, fg=self.origFG)
        ede.delete(0, 'end')
        ede.configure(bg=self.origBG, fg=self.origFG)

    def setEntry(self, she, sde, ede):
        she = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
        sde = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
        ede = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

    # TO DO:  COMPACT CODE
    def processEntryWidgets(self, flag):

        if flag.upper() == "SET":
            # ENTRY FIELDS

            # self.setEntry(self.startHeaderEntry1, self.startDataEntry1, self.endDataEntry1)
            # self.setEntry(self.startHeaderEntry2, self.startDataEntry2, self.endDataEntry2)
            # self.setEntry(self.startHeaderEntry3, self.startDataEntry3, self.endDataEntry3)
            # self.setEntry(self.startHeaderEntry4, self.startDataEntry4, self.endDataEntry4)
            # self.setEntry(self.startHeaderEntry5, self.startDataEntry5, self.endDataEntry5)

            self.titleHeader = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

            self.startHeaderEntry1 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.startDataEntry1 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.endDataEntry1 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

            self.startHeaderEntry2 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.startDataEntry2 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.endDataEntry2 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

            self.startHeaderEntry3 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.startDataEntry3 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.endDataEntry3 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

            self.startHeaderEntry4 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.startDataEntry4 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.endDataEntry4 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')

            self.startHeaderEntry5 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.startDataEntry5 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')
            self.endDataEntry5 = tk.Entry(self.frame2, width=20, font=(None, 12, 'bold'), fg='blue')


            self.titleHeader.focus()
            self.titleHeader.grid(row=2, column=2, padx=10, pady=5)

            self.startHeaderEntry1.grid(row=11, column=1, padx=10, pady=5)
            self.startDataEntry1.grid(row=11, column=2, padx=10, pady=5)
            self.endDataEntry1.grid(row=11, column=3, padx=10, pady=5)

            self.startHeaderEntry2.grid(row=12, column=1, padx=10, pady=5)
            self.startDataEntry2.grid(row=12, column=2, padx=10, pady=5)
            self.endDataEntry2.grid(row=12, column=3, padx=10, pady=5)

            self.startHeaderEntry3.grid(row=13, column=1, padx=10, pady=5)
            self.startDataEntry3.grid(row=13, column=2, padx=10, pady=5)
            self.endDataEntry3.grid(row=13, column=3, padx=10, pady=5)

            self.startHeaderEntry4.grid(row=14, column=1, padx=10, pady=5)
            self.startDataEntry4.grid(row=14, column=2, padx=10, pady=5)
            self.endDataEntry4.grid(row=14, column=3, padx=10, pady=5)

            self.startHeaderEntry5.grid(row=15, column=1, padx=10, pady=5)
            self.startDataEntry5.grid(row=15, column=2, padx=10, pady=5)
            self.endDataEntry5.grid(row=15, column=3, padx=10, pady=5)

            self.origBG = self.startHeaderEntry1.cget('bg')
            self.origFG = self.startHeaderEntry1.cget('fg')

        elif flag.upper() == "CLEAR":

            self.titleHeader.delete(0, 'end')
            self.titleHeader.configure(bg=self.origBG, fg=self.origFG)

            self.startHeaderEntry1.focus_set()

            self.clearEntry(self.startHeaderEntry1, self.startDataEntry1, self.endDataEntry1)
            self.clearEntry(self.startHeaderEntry2, self.startDataEntry2, self.endDataEntry2)
            self.clearEntry(self.startHeaderEntry3, self.startDataEntry3, self.endDataEntry3)
            self.clearEntry(self.startHeaderEntry4, self.startDataEntry4, self.endDataEntry4)
            self.clearEntry(self.startHeaderEntry5, self.startDataEntry5, self.endDataEntry5)

    def buttonState(self, action, sub, con, prv, nxt):
        if action.upper() == 'N':
            if sub:
                self.submitButton["state"] = "normal"
            if con:
                self.convertButton["state"] = "normal"
            if prv:
                self.prevButton["state"] = "normal"
            if nxt:
                self.nextButton["state"] = "normal"

        if action.upper() == 'D':
            if sub:
                self.submitButton["state"] = "disabled"
                self.submitButton["bg"] = "#C2C5CC"
            if con:
                self.convertButton["state"] = "disabled"
            if prv:
                self.prevButton["state"] = "disabled"
            if nxt:
                self.nextButton["state"] = "disabled"

    def getValidEntry(self, widget):
        if widget.upper() == 'E1':
            sheWidget = self.startHeaderEntry1
            sdeWidget = self.startDataEntry1
            edeWidget = self.endDataEntry1
        elif widget.upper() == 'E2':
            sheWidget = self.startHeaderEntry2
            sdeWidget = self.startDataEntry2
            edeWidget = self.endDataEntry2
        elif widget.upper() == 'E3':
            sheWidget = self.startHeaderEntry3
            sdeWidget = self.startDataEntry3
            edeWidget = self.endDataEntry3
        elif widget.upper() == 'E4':
            sheWidget = self.startHeaderEntry4
            sdeWidget = self.startDataEntry4
            edeWidget = self.endDataEntry4
        elif widget.upper() == 'E5':
            sheWidget = self.startHeaderEntry5
            sdeWidget = self.startDataEntry5
            edeWidget = self.endDataEntry5

        she = isValid(sheWidget.get())
        sde = isValid(sdeWidget.get())
        ede = isValid(edeWidget.get())

        if ede:
            edeWidget.configure(bg=self.origBG, fg=self.origFG)
        else:
            edeWidget.focus_set()
            edeWidget.configure(bg='#ffffe0', fg='red')

        if sde:
            sdeWidget.configure(bg=self.origBG, fg=self.origFG)
        else:
            sdeWidget.focus_set()
            sdeWidget.configure(bg='#ffffe0', fg='red')

        if she:
            sheWidget.configure(bg=self.origBG, fg=self.origFG)
        else:
            sheWidget.focus_set()
            sheWidget.configure(bg='#ffffe0', fg='red')

        if she and sde and ede:
            return {'beginHeader': she, 'beginData': sde, 'endData': ede}
        else:
            return None


# TO DO:  CAN SEPARATE THESE INTO ANOTHER FILE AS A MODULE
# INTEGRATION:  NEED CLEAN UP ON REDUNDANT FUNCTIONS
# ####################### FUNCTIONS #######################
def startConversionProcess(sh, tableList):
    print("tableList", tableList)

    # text = Properties(altText="SAMPLE")

    newMergedRangeList = unMerge(sh)
    print(newMergedRangeList)

    # accumulates all deleted rows
    totalEmptyRowList = []
    newTableList = []
    index = 0

    for sheetRange in tableList:
        print("SR: ", sheetRange)
        emptyRowList = []
        nnb = True
        if len(tableList) > 1:
            nnb = False
            # checks if there is no adjacent table (left and right)
            if index == 0:
                if column_index_from_string(tableList[index]['endData'][0]) > column_index_from_string(
                        tableList[index + 1]['beginData'][0]):
                    nnb = True
            elif index == len(tableList) - 1:
                if column_index_from_string(tableList[index]['beginData'][0]) < column_index_from_string(
                        tableList[index - 1]['endData'][0]):
                    nnb = True
            else:
                if column_index_from_string(tableList[index]['beginData'][0]) < column_index_from_string(
                        tableList[index - 1]['endData'][0]) \
                        and column_index_from_string(tableList[index]['endData'][0]) > column_index_from_string(
                        tableList[index + 1]['beginData'][0]):
                    nnb = True

        # checks if there are side by side tables
        # TO DO:  tables side by side
        if nnb:
            print("NO NEIGHBOUR")
            # adjusts by previous table
            tempSheetRange = getNewSheetRange('TR', sheetRange, emptyRowList, totalEmptyRowList)

            # returns list of empty rows while writing "End of Column" appropriately
            emptyRowList = checkEmptyRows(sh, tempSheetRange)
            print("emptyRows: ", emptyRowList)
            # adjusts by own rows
            newSheetRange = getNewSheetRange('NR', tempSheetRange, emptyRowList, totalEmptyRowList)
            print("nSR: ", sheetRange)
            # deletes specified empty rows
            deleteEmptyRows(sh, emptyRowList)

            newTableList.append(newSheetRange)
        else:
            pass

        totalEmptyRowList += emptyRowList

        # copies merged cell value across the merge range and updates Merge List
        reducedMergedRangeList = populateGroups(sh, newSheetRange, newMergedRangeList, emptyRowList)
        newMergedRangeList = reducedMergedRangeList

        table_count = len(tableList)

        # formats table into a Table format with the correct header
        convertTable(sh, newSheetRange, table_count)

        # formats whole table with borders
        start = timeit.default_timer()
        setBorder(sh, newSheetRange)
        stop = timeit.default_timer()
        print('Time: ', stop - start)

        index += 1

    # updates Screen Reader messages
    insertScreenReader(sh, newTableList)
    writeNoData(sh, newTableList)

    print("TO DO: Check overlapping table sheet entry")


# UN-MERGES MERGED CELLS IN THE SHEET
def unMerge(sh):
    # detects all merge cells and get the merge range
    mergedRangeList = sh.merged_cells.ranges
    # counts number of merged cells
    mergeCount = len(mergedRangeList)
    # stores the mergeRangeList before it gets unmerged
    tempMergedRangeList = []

    # un-merges merged cells
    counter = mergeCount
    while counter > 0:
        tempMergedRangeList.append(mergedRangeList[counter - 1])
        sh.unmerge_cells(str(mergedRangeList[counter - 1]))
        counter -= 1

    return tempMergedRangeList


# returns a list of dictionary
def getNewSheetRange(flag, sheetRange, emptyRowList, totalEmptyRowList):
    # determines new Data Range after deleting rows
    if flag.upper() == 'TR':
        beginHRow = sheetRange['beginHeader'][1] - len(totalEmptyRowList)
        beginDRow = sheetRange['beginData'][1] - len(totalEmptyRowList)
        endRow = sheetRange['endData'][1] - (len(emptyRowList) + len(totalEmptyRowList))
    elif flag.upper() == 'NR':
        beginHRow = sheetRange['beginHeader'][1]
        beginDRow = sheetRange['beginData'][1]
        for x in emptyRowList:
            # print(x, beginDRow)
            if x < sheetRange['beginData'][1]:
                beginDRow -= 1
        endRow = sheetRange['endData'][1] - len(emptyRowList)

    return {'beginHeader': (sheetRange['beginHeader'][0], beginHRow),
            'beginData': (sheetRange['beginData'][0], beginDRow), 'endData': (sheetRange['endData'][0], endRow)}


# RETURNS A DICTIONARY OF COORDINATES
def getPerimeter(sheetRange):
    beginCol = column_index_from_string(sheetRange['beginHeader'][0])
    beginRow = sheetRange['beginHeader'][1]
    endCol = column_index_from_string(sheetRange['endData'][0])
    endRow = sheetRange['endData'][1]

    return {'beginCol': beginCol, 'endCol': endCol, 'beginRow': beginRow, 'endRow': endRow}


# TEMPORARY:  only put EOC for each table
def checkEmptyRows(sh, sheetRange):
    emptyRowList = []
    # only used to shortened referenced names
    p = getPerimeter(sheetRange)

    # returns empty rows from Data Segment & writes "End of Column"
    for rowNum in range(p['beginRow'], p['endRow'] + 1):
        isEmpty = True
        for colNum in range(p['beginCol'], p['endCol'] + 1):
            if sh.cell(rowNum, colNum).value not in [None, '']:
                isEmpty = False
            if colNum + 1 == p['endCol'] + 1:
                sh.cell(rowNum, colNum + 1).value = "End Column"
                sh.cell(rowNum, colNum + 1).font = Font(color="FFFFFF")
        if isEmpty:
            emptyRowList.append(rowNum)

    return emptyRowList


# input: string => "B4"
# output: int column, row => (2, 4)
def getRowCol(cellAddress):
    try:
        column, row = coordinate_from_string(cellAddress)
        return column_index_from_string(column), row
    except:
        mb.showinfo("Alert", "Please enter a valid cell address (Letter(s) followed by Number(s): 'A67'")
        print("Please enter a valid cell address (Letter(s) followed by Number(s): 'A67'")


def deleteEmptyRows(sh, emptyRowList):
    deleted = 0
    # for row in reversed(emptyRowList):
    for row in emptyRowList:
        sh.delete_rows(row - deleted, 1)
        deleted += 1


# populate merged value across the group, adjusting for deleted rows
def populateGroups(sh, sheetRange, mergedRangeList, emptyRowList):
    tempMergeRangeList = []
    mergeCount = len(mergedRangeList)

    beginSheetColumn = column_index_from_string(sheetRange['beginHeader'][0])
    beginSheetRow = sheetRange['beginHeader'][1]
    endSheetColumn = column_index_from_string(sheetRange['endData'][0])
    endSheetRow = sheetRange['endData'][1]

    while mergeCount > 0:
        tempCellRange = mergedRangeList[mergeCount - 1]

        # gets the beginning and ending cells of the current merged cell
        mergedRange = str(tempCellRange).split(":")

        # gets the beginning cell of the current merged range
        cellBegin = mergedRange[0]
        # gets the coordinates of the beginning cell
        beginMergedColumn, beginMergedRow = getRowCol(cellBegin)

        # gets the ending cell of the current merged range
        cellEnd = mergedRange[1]
        # gets the coordinates of the ending cell
        endMergedColumn, endMergedRow = getRowCol(cellEnd)

        emptyRowList.sort()
        # adjust merged addressed by deleted rows
        shiftBy = 0
        for x in emptyRowList:
            # if merged cells includes deleted row
            # shift end of merged cells by number of rows deleted within the range
            if beginMergedRow < x <= endMergedRow:
                endMergedRow -= 1
            # if merged cells above deleted row
            # shifts beginning of merged cells by number of rows deleted above
            if x < beginMergedRow:
                shiftBy += 1

        beginMergedRow -= shiftBy
        endMergedRow -= shiftBy

        beginCell = "".join((get_column_letter(beginMergedColumn), str(beginMergedRow)))
        endCell = "".join((get_column_letter(endMergedColumn), str(endMergedRow)))

        newMergeRange = beginCell + ':' + endCell

        # manually checks if merged cell is subset of sheet range
        if beginSheetColumn <= beginMergedColumn <= endSheetColumn and beginSheetColumn <= endMergedColumn <= endSheetColumn \
                and beginSheetRow <= beginMergedRow < endSheetRow and beginSheetRow <= endMergedRow <= endSheetRow:

            # locates the current cell where the value for the merged cell is located
            currentCell = sh.cell(beginMergedRow, beginMergedColumn)

            # if merged cells in one row only => copy column across
            if beginMergedColumn != endMergedColumn and beginMergedRow == endMergedRow:
                for col in range(beginMergedColumn + 1, endMergedColumn + 1):
                    copyMergedValues(sh, beginMergedRow, col, currentCell)
            # if merged cells in one column only => copy row down
            elif beginMergedColumn == endMergedColumn and beginMergedRow != endMergedRow:
                for row in range(beginMergedRow + 1, endMergedRow + 1):
                    copyMergedValues(sh, row, beginMergedColumn, currentCell)
            # if merged cells spans multiple rows and columns
            else:
                for row in range(beginMergedRow, endMergedRow + 1):
                    for col in range(beginMergedColumn, endMergedColumn + 1):
                        copyMergedValues(sh, row, col, currentCell)

            # tempMergeRangeList.remove(tempCellRange)
        else:
            tempMergeRangeList.append(newMergeRange)

        mergeCount -= 1

    return tempMergeRangeList


def copyMergedValues(sh, row, col, currentCell):
    sh.cell(row, col).value = str(currentCell.value)
    sh.cell(row, col).font = copy(currentCell.font)
    sh.cell(row, col).fill = copy(currentCell.fill)
    sh.cell(row, col).border = copy(currentCell.border)
    sh.cell(row, col).alignment = copy(currentCell.alignment)


# INSERTS SCREEN READER TO IDENTIFY HEADERS AND DATA SEGMENT
def insertScreenReader(sh, tableList):
    blank = "Blank cells in the middle of the table represent No Data. "
    table_size = len(tableList)
    if table_size == 1:
        dataColumn = column_index_from_string(tableList[0]['beginData'][0])
        headerRow = tableList[0]['beginData'][1] - 1
        dataCell = "".join((get_column_letter(dataColumn), str(headerRow)))
        text = blank + "The table starts on " + dataCell + ".  Column Titles are in Row " + str(headerRow) + "."
    else:
        statement = ''
        for x in tableList:
            dataColumn = (x['beginData'][0])
            headerRow = x['beginData'][1] - 1
            dataCell = "".join((dataColumn, str(headerRow)))
            headers = ', ' + dataCell + ', Column Title in Row ' + str(headerRow)
            statement = statement + headers

        text = blank + 'There are ' + str(table_size) + ' tables in this sheet. The tables start at' + statement + '.'

    sh.cell(2, 1).value = text
    # sh.cell(2, 1).font = Font(color="0000FF")
    # a1 = sh['A1']
    # sh.cell(2, 1).font = copy(a1.font)


def convertTable(sh, sheetRange, count):
    display = "".join(('Table', str(count)))

    beginCell = "".join((sheetRange['beginData'][0], str(sheetRange['beginData'][1] - 1)))
    endCell = "".join((sheetRange['endData'][0], str(sheetRange['endData'][1])))

    dataRange = beginCell + ':' + endCell

    # table names must be unique for each sheet
    tab = Table(displayName=display, ref=dataRange)
    # Add a default style with striped rows and banded columns
    # name="TableStyleMedium9" /Light /Dark
    # style = TableStyleInfo(name="TableStyleLight5", showFirstColumn=False,
    #                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # tab.tableStyleInfo = style
    sh.add_table(tab)


# TO DO:  FORMAT HEADER1
def setBorder(sh, sheetRange):
    # adjust before coming here
    beginCell = "".join((sheetRange['beginHeader'][0], str(sheetRange['beginHeader'][1])))
    endCell = "".join((sheetRange['endData'][0], str(sheetRange['endData'][1])))

    dataRange = beginCell + ':' + endCell

    rows = sh[dataRange]

    # formats headers with thick borders
    headerRow = sheetRange['beginData'][1] - sheetRange['beginHeader'][1]

    borderHeight = 25

    hair = Side(border_style='hair', color='d3d3d3')
    thick = Side(border_style='thick', color='a9a9a9')
    alignV = Alignment(vertical='center')
    # alignH = Alignment(horizontal='center')
    align = Alignment(vertical='center', horizontal='center')

    maxRow = len(rows)  # index of the last row
    # print("MaxRow", maxRow)
    for row, cells in enumerate(rows, 1):
        # print("PosRow, cells: ", row, cells)
        maxCol = len(cells)  # index of the last cell
        # print("MaxCol", maxCol)
        for col, cell in enumerate(cells, 1):
            # print("PosCol, cell: ",  col, cell)

            sh.row_dimensions[cell.row].height = borderHeight
            cell.alignment = alignV

            if col != 1:
                cell.alignment = align

            border = Border(
                left=hair,
                right=hair,
                top=hair,
                bottom=hair
            )

            # applies thick border on corners
            if col == 1:
                border.left = thick
            if col == maxCol:
                border.right = thick
            if row == 1:
                border.top = thick
            if row == maxRow:
                border.bottom = thick
            # applies thick border on header rows
            if row == headerRow:
                border.top = thick
                border.bottom = thick
                cell.font = Font(bold=True)

            cell.border = border

    sh.row_dimensions[sh.max_row + 1].height = borderHeight


def formatReportTitle(sh, title_address):
    cell = "".join((title_address[0], str(title_address[1])))
    print("Cell to format: ", cell)
    title = sh[cell]
    title.style = 'Headline 1'


# def getTableTitles(sh, tableList):
#     titleNames = []
#     index = 0
#     for sheetRange in tableList:
#         cell = "".join((sheetRange['titleHeader'][0], str(sheetRange['titleHeader'][1])))
#         title = sh[cell]
#         titleNames.append((index+1, title.value))
#
#     return titleNames


def writeNoData(sh, tableList):

    def writeND(br, er, ec):
        # writes "No Data" above Header/Data Segment & "End of Column"
        for row in range(br, er):
            for col in range(1, ec):
                if sh.cell(row, col).value in [None, '', 0]:
                    sh.cell(row, col).value = "No Data"
                    # ADDED ALL FONT COLORS TO GREEN FOR NOW
                    sh.cell(row, col).font = Font(color="FFFFFF")
                # determines what is endCol
                if col + 1 == endCol:
                    sh.cell(row, col + 1).value = "End Column"
                    sh.cell(row, col + 1).font = Font(color="FFFFFF")

    index = 0
    for sheetRange in tableList:
        endCol = column_index_from_string(sheetRange['endData'][0]) + 1
        endRow = sheetRange['beginHeader'][1]

        if index == 0:
            beginRow = 1
        else:
            beginRow = tableList[index - 1]['endData'][1] + 1

        writeND(beginRow, endRow, endCol)
        index += 1

    beginRow = tableList[len(tableList) - 1]['endData'][1] + 1
    endRow = sh.max_row
    endCol = column_index_from_string(tableList[len(tableList) - 1]['endData'][0]) + 1
    print('MAX ROW:', endRow)

    writeND(beginRow, endRow, endCol)


if __name__ == "__main__":
    main()
